"""
KLA 2026 Candidate Data Scraper
================================
Fetches the official 2026 Kerala Legislative Assembly Election candidate
dataset from OpenDataKerala / Zenodo and saves it as CSV and Excel,
ordered by constituency number (1–140).

Data source : https://zenodo.org/records/19323710
License     : ODbL (Open Database License)
Publisher   : OpenDataKerala Community
"""

import csv
import io
import re
import difflib
import unicodedata
import urllib.request
import json
import tempfile
import zipfile
import os

DATA_URL = (
    "https://zenodo.org/records/19323710/files/"
    "Kerala%20Legislative%20Assembly%20Election%202026%20Candidate%20Data.csv"
    "?download=1"
)

# Official seat order 1–140 with district, keyed by constituency name.
# strip() is applied on lookup to handle trailing spaces in source data.
CONSTITUENCY_INFO = {
    "Manjeshwaram":       (1,   "Kasaragod"),
    "Kasaragod":          (2,   "Kasaragod"),
    "Rajapuram":          (3,   "Kasaragod"),
    "Uduma":              (4,   "Kasaragod"),
    "Kanhangad":          (5,   "Kasaragod"),
    "Trikaripur":         (6,   "Kasaragod"),
    "Payyannur":          (7,   "Kannur"),
    "Kalliasseri":        (8,   "Kannur"),
    "Thalassery":         (9,   "Kannur"),
    "Kuthuparamba":       (10,  "Kannur"),
    "Mattannur":          (11,  "Kannur"),
    "Peravoor":           (12,  "Kannur"),
    "Kannur":             (13,  "Kannur"),
    "Dharmadom":          (14,  "Kannur"),
    "Thalipparamba":      (15,  "Kannur"),
    "Irikkur":            (16,  "Kannur"),
    "Azhikode":           (17,  "Kannur"),
    "Sreekandapuram":     (18,  "Kannur"),
    "Vatakara":           (19,  "Kozhikode"),
    "Nadapuram":          (20,  "Kozhikode"),
    "Koyilandy":          (21,  "Kozhikode"),
    "Perambra":           (22,  "Kozhikode"),
    "Balussery":          (23,  "Kozhikode"),
    "Elathur":            (24,  "Kozhikode"),
    "Kozhikode North":    (25,  "Kozhikode"),
    "Kozhikode South":    (26,  "Kozhikode"),
    "Beypore":            (27,  "Kozhikode"),
    "Kunnamangalam":      (28,  "Kozhikode"),
    "Koduvally":          (29,  "Kozhikode"),
    "Thiruvambady":       (30,  "Kozhikode"),
    "Kondotty":           (31,  "Malappuram"),
    "Manjeri":            (32,  "Malappuram"),
    "Malappuram":         (33,  "Malappuram"),
    "Vengara":            (34,  "Malappuram"),
    "Tirur":              (35,  "Malappuram"),
    "Tanur":              (36,  "Malappuram"),
    "Tirurangadi":        (37,  "Malappuram"),
    "Kottakkal":          (38,  "Malappuram"),
    "Trikkalangode":      (39,  "Malappuram"),
    "Wandoor":            (40,  "Malappuram"),
    "Perinthalmanna":     (41,  "Malappuram"),
    "Mannarkkad":         (42,  "Palakkad"),
    "Malampuzha":         (43,  "Palakkad"),
    "Palakkad":           (44,  "Palakkad"),
    "Thrithala":          (45,  "Palakkad"),
    "Pattambi":           (46,  "Palakkad"),
    "Shornur":            (47,  "Palakkad"),
    "Ottapalam":          (48,  "Palakkad"),
    "Kongad":             (49,  "Palakkad"),
    "Alathur":            (50,  "Palakkad"),
    "Chittur":            (51,  "Palakkad"),
    "Nenmara":            (52,  "Palakkad"),
    "Kollengode":         (53,  "Palakkad"),
    "Wadakkancherry":     (54,  "Thrissur"),
    "Kunnamkulam":        (55,  "Thrissur"),
    "Guruvayur":          (56,  "Thrissur"),
    "Manalur":            (57,  "Thrissur"),
    "Ollur":              (58,  "Thrissur"),
    "Thrissur":           (59,  "Thrissur"),
    "Nattika":            (60,  "Thrissur"),
    "Irinjalakuda":       (61,  "Thrissur"),
    "Puthukkad":          (62,  "Thrissur"),
    "Chalakudy":          (63,  "Thrissur"),
    "Kaipamangalam":      (64,  "Thrissur"),
    "Perumbavoor":        (65,  "Ernakulam"),
    "Angamaly":           (66,  "Ernakulam"),
    "Aluva":              (67,  "Ernakulam"),
    "Kalamassery":        (68,  "Ernakulam"),
    "Paravur":            (69,  "Ernakulam"),
    "Vypen":              (70,  "Ernakulam"),
    "Kochi":              (71,  "Ernakulam"),
    "Thrippunithura":     (72,  "Ernakulam"),
    "Ernakulam":          (73,  "Ernakulam"),
    "Kolam":              (74,  "Ernakulam"),
    "Kunnathunadu":       (75,  "Ernakulam"),
    "Kothamangalam":      (76,  "Ernakulam"),
    "Muvattupuzha":       (77,  "Ernakulam"),
    "Piravom":            (78,  "Ernakulam"),
    "Idukki":             (79,  "Idukki"),
    "Udumbanchola":       (80,  "Idukki"),
    "Thodupuzha":         (81,  "Idukki"),
    "Peerumade":          (82,  "Idukki"),
    "Devikulam":          (83,  "Idukki"),
    "Vaikom":             (84,  "Kottayam"),
    "Ettumanoor":         (85,  "Kottayam"),
    "Kottayam":           (86,  "Kottayam"),
    "Poonjar":            (87,  "Kottayam"),
    "Erattupetta":        (88,  "Kottayam"),
    "Kaduthuruthy":       (89,  "Kottayam"),
    "Changanacherry":     (90,  "Kottayam"),
    "Kuttanad":           (91,  "Alappuzha"),
    "Alappuzha":          (92,  "Alappuzha"),
    "Ambalappuzha":       (93,  "Alappuzha"),
    "Cherthala":          (94,  "Alappuzha"),
    "Mararikulam":        (95,  "Alappuzha"),
    "Aroor":              (96,  "Alappuzha"),
    "Chengannur":         (97,  "Alappuzha"),
    "Mavelikkara":        (98,  "Alappuzha"),
    "Kayamkulam":         (99,  "Alappuzha"),
    "Haripad":            (100, "Alappuzha"),
    "Pathanamthitta":     (101, "Pathanamthitta"),
    "Ranny":              (102, "Pathanamthitta"),
    "Aranmula":           (103, "Pathanamthitta"),
    "Thiruvalla":         (104, "Pathanamthitta"),
    "Adoor":              (105, "Pathanamthitta"),
    "Konni":              (106, "Pathanamthitta"),
    "Chadayamangalam":    (107, "Kollam"),
    "Punalur":            (108, "Kollam"),
    "Pathanapuram":       (109, "Kollam"),
    "Kottarakkara":       (110, "Kollam"),
    "Kunnathur":          (111, "Kollam"),
    "Kundara":            (112, "Kollam"),
    "Chavara":            (113, "Kollam"),
    "Eravipuram":         (114, "Kollam"),
    "Kollam":             (115, "Kollam"),
    "Parassala":          (116, "Thiruvananthapuram"),
    "Kattakada":          (117, "Thiruvananthapuram"),
    "Varkala":            (118, "Thiruvananthapuram"),
    "Attingal":           (119, "Thiruvananthapuram"),
    "Vamanapuram":        (120, "Thiruvananthapuram"),
    "Aruvikkara":         (121, "Thiruvananthapuram"),
    "Nedumangad":         (122, "Thiruvananthapuram"),
    "Vattiyoorkavu":      (123, "Thiruvananthapuram"),
    "Thiruvananthapuram": (124, "Thiruvananthapuram"),
    "Kazhakuttam":        (125, "Thiruvananthapuram"),
    "Nemom":              (126, "Thiruvananthapuram"),
}
# Normalized lookup (handles casing, punctuation and minor spelling variants)
def _normalize_key(s: str) -> str:
    s = unicodedata.normalize("NFKD", (s or "")).lower()
    s = re.sub(r"[^a-z0-9 ]+", "", s)
    return " ".join(s.split())


# Build normalized mapping for fast lookup (may be replaced at runtime)
CONSTITUENCY_INFO_NORM = { _normalize_key(k): v for k, v in CONSTITUENCY_INFO.items() }


DATA_ZIP_URL = "https://zenodo.org/records/19323427/files/KLA2026-main.zip"


def fetch_constituency_mapping():
    """Attempt to download the official KLA2026 site archive from Zenodo and
    extract the constituency list (number, name, district). Returns a dict
    keyed by canonical name -> (int(number), district).
    On failure, returns the existing `CONSTITUENCY_INFO` as fallback.
    """
    try:
        print("Fetching constituency mapping from Zenodo...")
        req = urllib.request.Request(DATA_ZIP_URL, headers={"User-Agent": "KLA2026-scraper/1.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = resp.read()

        # write to temp file and read zip
        with tempfile.NamedTemporaryFile(delete=False) as tmp:
            tmp.write(data)
            tmp_path = tmp.name

        mapping = {}
        with zipfile.ZipFile(tmp_path) as z:
            try:
                content = z.read('KLA2026-main/src/data/constituencies.json').decode('utf-8')
            except KeyError:
                # older archive layout
                content = z.read('src/data/constituencies.json').decode('utf-8')
            arr = json.loads(content)
            for item in arr:
                name = item.get('name')
                number = int(item.get('number') or 0)
                district = item.get('district') or ''
                mapping[name] = (number, district)

        try:
            os.remove(tmp_path)
        except Exception:
            pass
        print(f"Loaded {len(mapping)} constituencies from archive")
        return mapping
    except Exception as e:
        print("Failed to fetch constituency mapping:", e)
        print("Falling back to built-in mapping (may be incomplete)")
        return CONSTITUENCY_INFO

FIELDS = [
    "Constituency No.",
    "Constituency Name",
    "District",
    "Alliance",
    "Party",
    "Candidate Name",
]

ALLIANCE_ORDER = {"LDF": 0, "UDF": 1, "NDA": 2}


def fetch_raw_csv(url):
    print("Fetching data from Zenodo...")
    req = urllib.request.Request(url, headers={"User-Agent": "KLA2026-scraper/1.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read().decode("utf-8-sig")


def parse(raw_csv):
    reader = csv.DictReader(io.StringIO(raw_csv))
    rows = []
    for r in reader:
        name = r.get("constituencyName", "").strip()
        key = _normalize_key(name)
        canonical_name = name
        if key in CONSTITUENCY_INFO_NORM:
            seat_no, district = CONSTITUENCY_INFO_NORM[key]
            # restore canonical name from original dict
            # find the original key whose normalized form matches
            for orig in CONSTITUENCY_INFO:
                if _normalize_key(orig) == key:
                    canonical_name = orig
                    break
        else:
            # try fuzzy match to catch small spelling variants (lower cutoff)
            match = difflib.get_close_matches(key, CONSTITUENCY_INFO_NORM.keys(), n=1, cutoff=0.7)
            if match:
                seat_no, district = CONSTITUENCY_INFO_NORM[match[0]]
                # use canonical name from original mapping
                for orig in CONSTITUENCY_INFO:
                    if _normalize_key(orig) == match[0]:
                        canonical_name = orig
                        break
            else:
                seat_no, district = (9999, "Unknown")
        rows.append({
            "Constituency No.":  seat_no,
            "Constituency Name": canonical_name,
            "District":          district,
            "Alliance":          r.get("alliance", "").strip(),
            "Party":             r.get("party", "").strip(),
            "Candidate Name":    r.get("candidateName", "").strip(),
        })
    # Sort by seat number (1–140), then LDF → UDF → NDA → others within each seat
    rows.sort(key=lambda r: (
        r["Constituency No."],
        ALLIANCE_ORDER.get(r["Alliance"], 99)
    ))
    return rows


def save_csv(rows, path="KLA2026_Candidates.csv"):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    print(f"CSV  saved → {path}  ({len(rows)} rows)")


def save_excel(rows, path="KLA2026_Candidates.xlsx"):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl not installed — skipping Excel output.")
        print("Install with:  pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "KLA2026 Candidates"

    # Header row
    ws.append(FIELDS)
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1565C0")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row in rows:
        ws.append([row[f] for f in FIELDS])

    # Column widths
    for col, width in zip("ABCDEF", [16, 24, 18, 10, 22, 30]):
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"Excel saved → {path}  ({len(rows)} rows)")


def main():
    raw = fetch_raw_csv(DATA_URL)
    # refresh constituency mapping from the official site archive (if available)
    try:
        global CONSTITUENCY_INFO, CONSTITUENCY_INFO_NORM
        mapping = fetch_constituency_mapping()
        # mapping: name -> (number, district)
        # update/replace existing mapping entries
        CONSTITUENCY_INFO.update(mapping)
        CONSTITUENCY_INFO_NORM = { _normalize_key(k): v for k, v in CONSTITUENCY_INFO.items() }
    except Exception:
        # on any failure continue with the built-in mapping
        pass

    rows = parse(raw)
    constituencies = len({r["Constituency No."] for r in rows})
    print(f"Done — {len(rows)} candidates across {constituencies} constituencies\n")
    save_csv(rows)
    save_excel(rows)


if __name__ == "__main__":
    main()